"""
Recognized MD -> Tables Excel
─────────────────────────────────────────────────────────────────
Collects every table found in the recognized Markdown, in document order,
into ONE worksheet. Each table is preceded by a title row carrying the
section heading it belongs to (the nearest preceding Markdown heading,
shown as a breadcrumb like "Section 2 > Financial Highlights").

This is the "tables Excel" output of the pure-OCR feature — it contains the
recognized table content only, never extracted metrics (metric extraction
is a separate, explicit feature: scripts/extract_metrics.py).

Table sources handled:
  - HTML <table> blocks (rowspan/colspan expanded into a plain grid)
  - Markdown pipe tables

Dependency: openpyxl
Reuses the table-grid parsers from scripts/md_to_docx.py (no changes there).
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

# Windows GBK console compatibility: switch stdout/stderr to UTF-8
for _s in (sys.stdout, sys.stderr):
    if hasattr(_s, "reconfigure"):
        try:
            _s.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

_SCRIPT_DIR = Path(__file__).resolve().parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

# reuse the battle-tested grid parsers (existing module, imported untouched)
from md_to_docx import (  # noqa: E402
    _parse_html_table, _parse_pipe_table, _strip_inline_md,
    _HEADING_RE, _PIPE_ROW_RE, _SLICE_RE,
)


# ==============================================================================
# Table collection (document order, with owning-heading breadcrumb)
# ==============================================================================

def collect_tables(md_text: str) -> list[dict]:
    """
    Walk the Markdown top to bottom, tracking the heading stack, and collect
    every table as {"title": breadcrumb, "grid": [[cell, ...], ...]}.
    Physical-slice marker lines are ignored (they are wire format, not
    document headings).
    """
    lines = md_text.splitlines()
    heading_stack: list[tuple[int, str]] = []   # [(level, text), ...]
    tables: list[dict] = []
    i = 0

    def _breadcrumb() -> str:
        if not heading_stack:
            return "(no heading)"
        return " > ".join(t for _, t in heading_stack)

    while i < len(lines):
        s = lines[i].strip()

        if _SLICE_RE.match(s):                       # slice marker: not a heading
            i += 1
            continue

        m = _HEADING_RE.match(s)
        if m:
            level = min(len(m.group(1)), 6)
            text = _strip_inline_md(m.group(2))
            while heading_stack and heading_stack[-1][0] >= level:
                heading_stack.pop()
            heading_stack.append((level, text))
            i += 1
            continue

        low = s.lower()
        if low.startswith("<table"):                  # HTML table block
            block = [lines[i]]
            while "</table>" not in lines[i].lower() and i + 1 < len(lines):
                i += 1
                block.append(lines[i])
            grid = _parse_html_table("\n".join(block))
            if grid:
                tables.append({"title": _breadcrumb(), "grid": grid})
            i += 1
            continue

        if _PIPE_ROW_RE.match(s):                     # pipe table block
            block = []
            while i < len(lines) and _PIPE_ROW_RE.match(lines[i].strip()):
                block.append(lines[i])
                i += 1
            grid = _parse_pipe_table(block)
            if grid:
                tables.append({"title": _breadcrumb(), "grid": grid})
            continue

        i += 1

    return tables


# ==============================================================================
# Excel writing (one sheet, sequential, titled blocks)
# ==============================================================================

def _display_len(text: str) -> float:
    """Approximate display width (CJK chars count as 1.8)."""
    n = 0.0
    for ch in text:
        n += 1.8 if "一" <= ch <= "鿿" else 1.0
    return n


def write_tables_xlsx(tables: list[dict], out_path: str | Path,
                      source_name: str = "") -> Path:
    """Write all tables sequentially into one worksheet."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tables"

    title_font = Font(bold=True, size=11, color="1e5cb8")
    title_fill = PatternFill("solid", fgColor="e8f0fc")
    header_font = Font(bold=True, size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)

    col_widths: dict[int, float] = {}
    row = 1

    if source_name:
        ws.cell(row=row, column=1, value=f"Source: {source_name}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 2

    for n, tb in enumerate(tables, 1):
        grid = tb["grid"]
        width = max((len(r) for r in grid), default=0)
        if width == 0:
            continue

        # title row: "Table N · <owning heading breadcrumb>", merged across table
        tcell = ws.cell(row=row, column=1, value=f"Table {n} · {tb['title']}")
        tcell.font = title_font
        tcell.fill = title_fill
        if width > 1:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=width)
        row += 1

        # grid rows (first row treated as header)
        for ri, grow in enumerate(grid):
            for ci, val in enumerate(grow, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.alignment = cell_align
                if ri == 0:
                    c.font = header_font
                w = min(_display_len(str(val)), 60.0)
                if w > col_widths.get(ci, 0):
                    col_widths[ci] = w
            row += 1

        row += 1  # blank row between tables

    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = max(8.0, w + 2)

    wb.save(str(out_path))
    return out_path


def md_to_excel(md_path: str | Path, out_path: str | Path) -> tuple[Path, int]:
    """MD file -> tables Excel. Returns (output path, table count)."""
    md_path = Path(md_path)
    md_text = md_path.read_text(encoding="utf-8-sig", errors="replace")
    tables = collect_tables(md_text)
    out = write_tables_xlsx(tables, out_path, source_name=md_path.stem)
    print(f"[xlsx] ✅ {len(tables)} table(s) -> {out}")
    return out, len(tables)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description=("Collect every table from a recognized MD into one Excel "
                     "sheet, in document order, titled by owning section heading."))
    ap.add_argument("md", help="recognized MD path")
    ap.add_argument("output", help="output .xlsx path")
    args = ap.parse_args()
    md_to_excel(args.md, args.output)
