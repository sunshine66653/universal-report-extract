"""
Excel rule sheet -> JSON converter
─────────────────────────────────────────────────────────────────
Maintainers manage metric rules in Excel; this module converts them to the
engine's internal JSON at runtime.

Supported columns (headers are case-insensitive; missing columns default):
    id              metric id (required, unique)
    name            metric name (required)
    source          definition / where to take the number from (fed to the LLM)
    enabled         TRUE/FALSE/1/0/是/否 (default TRUE)
    section_hint    section locator keywords, comma/semicolon/newline separated;
                    "/" inside one token means OR
    aliases         alias terms, same separators; "/" inside one token means OR
    extraction_mode direct / calc (default direct)
    calc            calculation expression or JSON (when extraction_mode=calc)
    aggregation     aggregation method (optional)
    allow_parent_note  whether parent-company note sections may match
                       (TRUE/FALSE, default FALSE)
    skip_reason     remarks / skip reason (optional)

Multi-sheet: every sheet is one "rule group"; the sheet name becomes `group`.

NOTE: _COLUMN_ALIASES contains Chinese header names on purpose — real rule
sheets (e.g. profiles/cn_securities) use Chinese column headers. Keep them.
"""
from __future__ import annotations

import json
import hashlib
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


# ── header normalization map (Chinese + English headers both accepted) ──
_COLUMN_ALIASES = {
    "id": "id",
    "编号": "id",
    "name": "name",
    "名称": "name",
    "指标名称": "name",
    "name_in_material": "name_in_material",
    "素材名称": "name_in_material",
    "素材中的叫法": "name_in_material",
    "source": "source",
    "来源": "source",
    "来源说明": "source",
    "口径": "source",
    "enabled": "enabled",
    "启用": "enabled",
    "是否启用": "enabled",
    "section_hint": "section_hint",
    "章节线索": "section_hint",
    "定位线索": "section_hint",
    "aliases": "aliases",
    "别名": "aliases",
    "extraction_mode": "extraction_mode",
    "抽取模式": "extraction_mode",
    "calc": "calc",
    "计算": "calc",
    "aggregation": "aggregation",
    "聚合": "aggregation",
    "allow_parent_note": "allow_parent_note",
    "母公司附注": "allow_parent_note",
    "skip_reason": "skip_reason",
    "备注": "skip_reason",
    "跳过原因": "skip_reason",
}

_LIST_FIELDS = {"section_hint", "aliases"}
_BOOL_FIELDS = {"enabled", "allow_parent_note"}


def _norm_header(h: Any) -> Optional[str]:
    if h is None:
        return None
    key = str(h).strip().lower()
    return _COLUMN_ALIASES.get(key, _COLUMN_ALIASES.get(str(h).strip(), None))


def _parse_bool(v: Any, default: bool = True) -> bool:
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    # accepts Chinese 是 / check marks as truthy (real sheets use them)
    return s in ("true", "1", "yes", "y", "是", "√", "✓", "t")


def _parse_list(v: Any) -> List[str]:
    if v is None or v == "":
        return []
    s = str(v)
    # commas / semicolons / full-width commas / newlines all act as separators;
    # "/" is preserved (in-token OR)
    for sep in ["\n", "；", ";", "，"]:
        s = s.replace(sep, ",")
    return [x.strip() for x in s.split(",") if x.strip()]


def excel_to_rules(xlsx_path: str | Path) -> List[Dict[str, Any]]:
    """Read the Excel rule sheet, return a list of rules."""
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rules: List[Dict[str, Any]] = []

    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        # header row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_map: Dict[int, str] = {}
        for ci, h in enumerate(header_row):
            nh = _norm_header(h)
            if nh:
                col_map[ci] = nh
        if "id" not in col_map.values() or "name" not in col_map.values():
            # not a rule sheet — skip
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            rec: Dict[str, Any] = {}
            for ci, field in col_map.items():
                val = row[ci] if ci < len(row) else None
                if field in _LIST_FIELDS:
                    rec[field] = _parse_list(val)
                elif field in _BOOL_FIELDS:
                    default = True if field == "enabled" else False
                    rec[field] = _parse_bool(val, default)
                else:
                    rec[field] = (str(val).strip() if val is not None else None)
            if not rec.get("id") or not rec.get("name"):
                continue
            rec.setdefault("extraction_mode", "direct")
            rec.setdefault("enabled", True)
            rec["group"] = ws.title
            rules.append(rec)

    return rules


# canonical column order for generated sheets (headers are the English
# canonical names — excel_to_rules maps them back 1:1)
_EXPORT_COLUMNS = ["id", "name", "name_in_material", "source", "enabled",
                   "section_hint", "aliases", "extraction_mode", "calc",
                   "aggregation", "allow_parent_note", "skip_reason"]


def rules_to_excel(rules: List[Dict[str, Any]],
                   xlsx_path: str | Path) -> Path:
    """Inverse of excel_to_rules: materialize a hand-editable Excel from
    rules in JSON form (e.g. when the skill was distributed without the
    .xlsx). One sheet per `group`; list fields joined with ", ", booleans
    written as TRUE/FALSE. Round-trips losslessly through excel_to_rules."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for r in rules:
        groups.setdefault(str(r.get("group") or "rules"), []).append(r)

    used_titles: set = set()
    for gname, grules in groups.items():
        title = re.sub(r'[:\\/?*\[\]]', "_", gname).strip()[:31] or "rules"
        base, k = title, 1
        while title in used_titles:
            title = f"{base[:28]}_{k}"
            k += 1
        used_titles.add(title)
        ws = wb.create_sheet(title=title)
        ws.append(_EXPORT_COLUMNS)
        for r in grules:
            row = []
            for col in _EXPORT_COLUMNS:
                v = r.get(col)
                if col in _LIST_FIELDS:
                    row.append(", ".join(v) if v else None)
                elif col in _BOOL_FIELDS:
                    if v is None:
                        v = (col == "enabled")   # same defaults as the parser
                    row.append("TRUE" if v else "FALSE")
                else:
                    row.append(None if v in (None, "") else str(v))
            ws.append(row)

    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(xlsx_path))
    return xlsx_path


def _file_sig(path: Path) -> str:
    st = path.stat()
    raw = f"{path}|{st.st_mtime_ns}|{st.st_size}"
    return hashlib.md5(raw.encode()).hexdigest()[:12]


def excel_to_rules_json(
    xlsx_path: str | Path,
    json_path: Optional[str | Path] = None,
    force: bool = False,
) -> Path:
    """
    Convert and cache as JSON. The Excel is the single hand-maintained
    source of truth; the JSON is its cache, auto-regenerated whenever the
    Excel changes (signature = mtime+size). Idempotent otherwise.
    Returns the JSON path.
    """
    xlsx_path = Path(xlsx_path)
    if json_path is None:
        json_path = xlsx_path.with_suffix(".json")
    json_path = Path(json_path)

    sig = _file_sig(xlsx_path)
    if json_path.exists() and not force:
        try:
            cached = json.loads(json_path.read_text(encoding="utf-8"))
            if isinstance(cached, dict) and cached.get("_sig") == sig:
                return json_path
        except Exception:
            pass

    rules = excel_to_rules(xlsx_path)
    payload = {
        "_sig": sig,
        "_source": xlsx_path.name,
        "count": len(rules),
        "rules": rules,
    }
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8",
    )
    return json_path


def load_rules(
    rules_path: str | Path, force_convert: bool = False,
) -> List[Dict[str, Any]]:
    """
    Unified entry: a .xlsx path is converted to JSON first; a .json path is
    read directly. Returns enabled rules only.
    """
    rules_path = Path(rules_path)
    if rules_path.suffix.lower() in (".xlsx", ".xlsm"):
        json_path = excel_to_rules_json(rules_path, force=force_convert)
        data = json.loads(json_path.read_text(encoding="utf-8"))
        rules = data.get("rules", data) if isinstance(data, dict) else data
    else:
        data = json.loads(rules_path.read_text(encoding="utf-8"))
        rules = data.get("rules", data) if isinstance(data, dict) else data

    return [r for r in rules if r.get("enabled", True)]


def load_rules_raw(rules_path: str | Path) -> List[Dict[str, Any]]:
    """Read ALL rules (disabled included, no Excel->JSON conversion side
    effects) — for validation/tooling. Accepts .xlsx/.xlsm/.json; the JSON
    may be the cached wrapper ({"rules": [...]}) or a plain array."""
    p = Path(rules_path)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        return excel_to_rules(p)
    data = json.loads(p.read_text(encoding="utf-8"))
    return data.get("rules", data) if isinstance(data, dict) else data


def validate_rules(rules: List[Dict[str, Any]]) -> List[str]:
    """Lint a rule list (typically a hand-edited JSON). Returns a list of
    human-readable problems; empty list = clean."""
    problems: List[str] = []
    seen: Dict[Any, int] = {}
    if not isinstance(rules, list):
        return ["top level must be a JSON array of rule objects "
                '(or {"rules": [...]})']
    for i, r in enumerate(rules):
        tag = f"rule #{i + 1}"
        if not isinstance(r, dict):
            problems.append(f"{tag}: not an object")
            continue
        rid = r.get("id")
        tag = f"rule #{i + 1} (id={rid!r})"
        if not rid:
            problems.append(f"{tag}: missing id")
        elif rid in seen:
            problems.append(f"{tag}: duplicate id (first seen at "
                            f"rule #{seen[rid] + 1})")
        else:
            seen[rid] = i
        if not r.get("name"):
            problems.append(f"{tag}: missing name")
        # extraction_mode is free-text guidance interpolated into the prompt
        # (direct / calc / prefer_direct_else_formula / ...) — no closed set;
        # only the plain "calc" mode hard-requires a formula
        if (r.get("extraction_mode") or "").strip() == "calc" and not r.get("calc"):
            problems.append(f"{tag}: extraction_mode=calc but calc is empty")
        for f in _LIST_FIELDS:
            v = r.get(f)
            if v is not None and not isinstance(v, list):
                problems.append(f'{tag}: {f} must be a JSON array, e.g. '
                                f'["货币资金", "现金及存放"]')
        for f in _BOOL_FIELDS:
            v = r.get(f)
            if v is not None and not isinstance(v, bool):
                problems.append(f"{tag}: {f} must be true/false")
    return problems
